"""Excel formatting utilities for openpyxl workbooks."""

from __future__ import annotations

from datetime import datetime

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ars_analysis.pipeline.context import PipelineContext

# Style constants
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1E3D59", end_color="1E3D59", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
DATA_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1E3D59")
SUBTITLE_FONT = Font(name="Calibri", size=11, color="666666")
KPI_VALUE_FONT = Font(name="Calibri", size=16, bold=True, color="1E3D59")
KPI_LABEL_FONT = Font(name="Calibri", size=10, color="666666")


def format_headers(ws: Worksheet) -> None:
    """Apply header formatting to first row: bold, fill, freeze."""
    if ws.max_row is None or ws.max_row < 1:
        return

    for col_idx in range(1, (ws.max_column or 0) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"


def auto_column_width(ws: Worksheet, max_width: int = 40) -> None:
    """Set column widths based on content (capped at max_width)."""
    for col_idx in range(1, (ws.max_column or 0) + 1):
        max_len = 8  # minimum
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    max_len = max(max_len, min(cell_len + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def apply_data_borders(ws: Worksheet) -> None:
    """Apply thin borders to all data cells."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.font = DATA_FONT


def format_worksheet(ws: Worksheet) -> None:
    """Apply full formatting to a data worksheet."""
    format_headers(ws)
    apply_data_borders(ws)
    auto_column_width(ws)


def create_summary_sheet(wb, ctx: PipelineContext) -> None:
    """Create a Summary sheet with client info and KPI overview.

    Inserts at position 0 (first sheet in workbook).
    """
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="ARS Analysis Summary")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    # Client info
    info_rows = [
        ("Client", ctx.client.client_name),
        ("Client ID", ctx.client.client_id),
        ("Month", ctx.client.month),
        ("CSM", ctx.client.assigned_csm or "N/A"),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]

    for i, (label, value) in enumerate(info_rows, 3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=value).font = DATA_FONT

    # KPIs from ctx.results
    kpi_start = len(info_rows) + 5
    ws.cell(row=kpi_start, column=1, value="Key Metrics").font = Font(
        bold=True, size=12, color="1E3D59"
    )

    kpis = _extract_kpis(ctx)
    for i, (label, value) in enumerate(kpis, kpi_start + 1):
        ws.cell(row=i, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=i, column=2, value=value).font = KPI_VALUE_FONT

    # Slides summary
    slide_start = kpi_start + len(kpis) + 3
    ws.cell(row=slide_start, column=1, value="Analyses").font = Font(
        bold=True, size=12, color="1E3D59"
    )

    success_count = sum(1 for s in ctx.all_slides if getattr(s, "success", True))
    ws.cell(row=slide_start + 1, column=1, value="Total Slides").font = KPI_LABEL_FONT
    ws.cell(row=slide_start + 1, column=2, value=len(ctx.all_slides)).font = KPI_VALUE_FONT
    ws.cell(row=slide_start + 2, column=1, value="Successful").font = KPI_LABEL_FONT
    ws.cell(row=slide_start + 2, column=2, value=success_count).font = KPI_VALUE_FONT

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20


def _extract_kpis(ctx: PipelineContext) -> list[tuple[str, str]]:
    """Pull key metrics from ctx.results for the summary sheet."""
    kpis: list[tuple[str, str]] = []

    # DCTR
    dctr_1 = ctx.results.get("dctr_1") or ctx.results.get("dctr.penetration", [])
    if isinstance(dctr_1, dict):
        ins = dctr_1.get("insights", {})
        if ins.get("overall_dctr"):
            kpis.append(("Overall DCTR", f"{ins['overall_dctr']:.1%}"))

    # Reg E
    rege_1 = ctx.results.get("reg_e_1", {})
    if isinstance(rege_1, dict) and rege_1.get("opt_in_rate"):
        kpis.append(("Reg E Opt-In Rate", f"{rege_1['opt_in_rate']:.1%}"))

    # Attrition
    att_1 = ctx.results.get("attrition_1", {})
    if isinstance(att_1, dict) and att_1.get("overall_rate"):
        kpis.append(("Attrition Rate", f"{att_1['overall_rate']:.1%}"))

    # Value
    val_1 = ctx.results.get("value_1", {})
    if isinstance(val_1, dict) and val_1.get("delta"):
        kpis.append(("IC Revenue Delta", f"${val_1['delta']:,.2f}"))

    # TXN
    txn = ctx.results.get("txn_summary", {})
    if isinstance(txn, dict) and txn.get("transaction_count"):
        kpis.append(("Transactions Analyzed", f"{txn['transaction_count']:,}"))

    return kpis
