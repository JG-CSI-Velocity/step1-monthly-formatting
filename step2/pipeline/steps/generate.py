"""Step: Generate deliverables (Excel, PowerPoint, archive)."""

from __future__ import annotations

import json
import shutil
from dataclasses import asdict, dataclass

import openpyxl
from loguru import logger

from ars_analysis.output.excel_formatter import (
    create_summary_sheet,
    format_worksheet,
)
from ars_analysis.pipeline.context import PipelineContext


@dataclass
class SlideStatus:
    """Diagnostic status for a single slide in the run report."""

    slide_id: str
    module_id: str
    success: bool
    has_chart: bool
    has_excel: bool
    error: str = ""
    title: str = ""


def _build_run_report(ctx: PipelineContext) -> list[SlideStatus]:
    """Build a diagnostic report of all slide statuses after analysis."""
    report: list[SlideStatus] = []
    for result in ctx.all_slides:
        module_id = ""
        # Determine module_id from ctx.results mapping
        for mid, results_list in ctx.results.items():
            if isinstance(results_list, list) and result in results_list:
                module_id = mid
                break

        report.append(
            SlideStatus(
                slide_id=result.slide_id,
                module_id=module_id,
                success=result.success,
                has_chart=bool(result.chart_path and result.chart_path.exists()),
                has_excel=bool(result.excel_data),
                error=result.error or "",
                title=result.title,
            )
        )
    return report


def _save_run_report(ctx: PipelineContext, report: list[SlideStatus]) -> None:
    """Save run report as JSON next to output files."""
    report_path = ctx.paths.base_dir / f"{ctx.client.client_id}_{ctx.client.month}_run_report.json"
    ctx.paths.base_dir.mkdir(parents=True, exist_ok=True)

    ok = sum(1 for s in report if s.success and s.has_chart)
    failed = sum(1 for s in report if not s.success)
    no_chart = sum(1 for s in report if s.success and not s.has_chart)

    report_data = {
        "client_id": ctx.client.client_id,
        "month": ctx.client.month,
        "summary": {
            "total": len(report),
            "ok": ok,
            "failed": failed,
            "no_chart": no_chart,
        },
        "slides": [asdict(s) for s in report],
    }
    report_path.write_text(json.dumps(report_data, indent=2))
    ctx.export_log.append(str(report_path))
    logger.info(
        "Run report: {ok}/{total} slides OK, {failed} failed, {no_chart} no chart",
        ok=ok,
        total=len(report),
        failed=failed,
        no_chart=no_chart,
    )


def step_generate(ctx: PipelineContext) -> None:
    """Generate all output deliverables from analysis results.

    Order: run report -> Excel workbook -> PowerPoint deck -> archive copy.
    Uses single-write pattern: build Excel once, then shutil.copy2 for master.
    """
    if not ctx.all_slides:
        logger.warning("No analysis results to generate deliverables from")
        return

    # Build and save run report before deck build (diagnostics first)
    report = _build_run_report(ctx)
    _save_run_report(ctx, report)
    ctx.results["_run_report"] = report

    _write_excel(ctx)
    _build_deck(ctx)
    logger.info("Deliverables generated for {client}", client=ctx.client.client_id)


def step_archive(ctx: PipelineContext) -> None:
    """Copy deliverables to archive location. Non-critical step."""
    logger.info("Archive step for {client} (not yet implemented)", client=ctx.client.client_id)


def _write_excel(ctx: PipelineContext) -> None:
    """Write all analysis results to a formatted Excel workbook.

    Single-write pattern: one workbook with a tab per analysis.
    Then shutil.copy2 for the master/archive copy.
    """
    excel_path = ctx.paths.excel_dir / f"{ctx.client.client_id}_{ctx.client.month}_analysis.xlsx"
    ctx.paths.excel_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    sheets_written = 0
    for result in ctx.all_slides:
        if result.excel_data is None:
            continue
        for sheet_name, df in result.excel_data.items():
            # Truncate sheet name to Excel 31-char limit
            safe_name = f"{result.slide_id}_{sheet_name}"[:31]
            ws = wb.create_sheet(title=safe_name)
            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            # Write data rows
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            # Format the worksheet
            format_worksheet(ws)
            sheets_written += 1

    if sheets_written == 0:
        logger.warning("No Excel data to write")
        return

    # Add summary sheet at the front
    create_summary_sheet(wb, ctx)

    wb.save(excel_path)
    ctx.export_log.append(str(excel_path))
    logger.info("Excel written: {path} ({n} sheets)", path=excel_path.name, n=sheets_written)

    # Single-write pattern: copy to master location if configured
    if ctx.settings and hasattr(ctx.settings, "paths"):
        master_dir = getattr(ctx.settings.paths, "ars_base", None)
        if master_dir and master_dir != ctx.paths.excel_dir:
            master_path = master_dir / excel_path.name
            try:
                shutil.copy2(excel_path, master_path)
                logger.info("Master copy: {name}", name=master_path.name)
            except OSError as exc:
                logger.warning("Master copy failed: {err}", err=exc)


def _build_deck(ctx: PipelineContext) -> None:
    """Build PowerPoint deck from analysis results."""
    skip_pptx = False
    if ctx.settings and hasattr(ctx.settings, "pipeline"):
        skip_pptx = getattr(ctx.settings.pipeline, "skip_pptx", False)

    if skip_pptx:
        logger.info("PowerPoint generation skipped (skip_pptx=True)")
        return

    try:
        from ars_analysis.output.deck_builder import build_deck

        result = build_deck(ctx)
        if result:
            if ctx.progress_callback:
                ctx.progress_callback(f"PPTX deck: {result.name}")
        else:
            msg = "PPTX: no slides with charts to build deck"
            logger.warning(msg)
            if ctx.progress_callback:
                ctx.progress_callback(msg)
    except ImportError:
        msg = f"PPTX skipped: deck_builder not available ({len(ctx.all_slides)} slides ready)"
        logger.warning(msg)
        if ctx.progress_callback:
            ctx.progress_callback(msg)
    except Exception as exc:
        msg = f"PPTX build failed: {exc}"
        logger.error(msg)
        ctx.export_log.append(f"ERROR: {msg}")
        if ctx.progress_callback:
            ctx.progress_callback(msg)
