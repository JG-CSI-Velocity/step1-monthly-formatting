"""7-step ODDD formatting pipeline + format validation checks.

Ported from ars-pipeline/src/ars/pipeline/steps/format.py (339 lines).
Transforms raw ODDD files into the formatted structure needed by ARS and ICS.

Steps:
    2. Drop PYTD/YTD columns
    3. Calculate totals, monthly averages, swipe categories
    4. Combine PIN+Sig into per-month Spend/Swipes
    5. Age calculations (DOB -> holder age, Date Opened -> account age)
    6. Mail & Response grouping (# of Offers, # of Responses, Response Grouping)
    7. Control segmentation (per-month Segmentation: Control/Responder/Non-Responder)
"""

from __future__ import annotations

import csv
import logging
import re
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)

_MONTH_MAP = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dec": 12,
}
_MONTH_PREFIX_RE = re.compile(r"^([A-Z][a-z]{2})(\d{2})")


def _infer_report_date(df: pd.DataFrame) -> pd.Timestamp:
    """Infer a stable report date from column headers or date fields."""
    month_tags: list[pd.Timestamp] = []
    for col in df.columns:
        match = _MONTH_PREFIX_RE.match(str(col))
        if not match:
            continue
        month = _MONTH_MAP.get(match.group(1))
        if not month:
            continue
        year = 2000 + int(match.group(2))
        month_tags.append(pd.Timestamp(year, month, 1))

    if month_tags:
        return max(month_tags)

    for col in ("Date Closed", "Date Opened"):
        if col in df.columns:
            parsed = pd.to_datetime(df[col], errors="coerce", format="mixed")
            max_date = parsed.max()
            if pd.notna(max_date):
                return pd.Timestamp(max_date).normalize()

    return pd.Timestamp.now()


def format_odd(df: pd.DataFrame) -> pd.DataFrame:
    """Apply all formatting steps to a raw ODDD DataFrame."""
    df = df.copy()
    df = _step2_drop_pytd_ytd(df)
    df = _step3_totals_averages_categories(df)
    df = _step4_combine_pin_sig(df)
    df = _step5_age_calculations(df)
    df = _step6_mail_response_grouping(df)
    df = _step7_control_segmentation(df)
    return df


def _step2_drop_pytd_ytd(df: pd.DataFrame) -> pd.DataFrame:
    """Remove prior-year-to-date and year-to-date columns."""
    drop_cols = [c for c in df.columns if "PYTD" in c or "YTD" in c]
    return df.drop(columns=drop_cols, errors="ignore")


def _step3_totals_averages_categories(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate totals, monthly averages, and swipe categories."""
    pin_spend = [c for c in df.columns if c.endswith("PIN $")]
    sig_spend = [c for c in df.columns if c.endswith("Sig $")]
    pin_count = [c for c in df.columns if c.endswith("PIN #")]
    sig_count = [c for c in df.columns if c.endswith("Sig #")]
    mtd_cols = [c for c in df.columns if c.endswith("MTD")]

    for col_list in [pin_spend, sig_spend, pin_count, sig_count, mtd_cols]:
        for c in col_list:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    df["Total Spend"] = df[pin_spend].sum(axis=1) + df[sig_spend].sum(axis=1)
    df["Total Swipes"] = df[pin_count].sum(axis=1) + df[sig_count].sum(axis=1)
    df["Total Items"] = df[mtd_cols].sum(axis=1) if mtd_cols else 0

    # Last 3 and 12 month sums
    for n, label in [(3, "3"), (12, "12")]:
        recent_pin_s = pin_spend[-n:] if len(pin_spend) >= n else pin_spend
        recent_sig_s = sig_spend[-n:] if len(sig_spend) >= n else sig_spend
        recent_pin_c = pin_count[-n:] if len(pin_count) >= n else pin_count
        recent_sig_c = sig_count[-n:] if len(sig_count) >= n else sig_count
        recent_mtd = mtd_cols[-n:] if len(mtd_cols) >= n else mtd_cols

        df[f"last {label}-mon spend"] = df[recent_pin_s].sum(axis=1) + df[recent_sig_s].sum(axis=1)
        df[f"last {label}-mon swipes"] = df[recent_pin_c].sum(axis=1) + df[recent_sig_c].sum(axis=1)
        if recent_mtd:
            df[f"Last {label}-mon Items"] = df[recent_mtd].sum(axis=1)
        else:
            df[f"Last {label}-mon Items"] = 0

    # Monthly averages
    for period, divisor in [("12", 12), ("3", 3)]:
        spend_key = f"last {period}-mon spend"
        swipe_key = f"last {period}-mon swipes"
        items_key = f"Last {period}-mon Items"

        df[f"MonthlySpend{period}"] = df[spend_key] / divisor
        df[f"MonthlySwipes{period}"] = df[swipe_key] / divisor
        df[f"MonthlyItems{period}"] = df[items_key] / divisor

    # Swipe categories
    df["SwipeCat12"] = df["MonthlySwipes12"].apply(_categorize_swipes)
    df["SwipeCat3"] = df["MonthlySwipes3"].apply(_categorize_swipes)

    return df


def _step4_combine_pin_sig(df: pd.DataFrame) -> pd.DataFrame:
    """Create combined Spend and Swipes columns for each month."""
    pin_spend_cols = [c for c in df.columns if c.endswith("PIN $")]
    for col in pin_spend_cols:
        prefix = col.replace(" PIN $", "")
        sig_col = f"{prefix} Sig $"
        if sig_col in df.columns:
            df[f"{prefix} Spend"] = df[col] + df[sig_col]

    pin_count_cols = [c for c in df.columns if c.endswith("PIN #")]
    for col in pin_count_cols:
        prefix = col.replace(" PIN #", "")
        sig_col = f"{prefix} Sig #"
        if sig_col in df.columns:
            df[f"{prefix} Swipes"] = df[col] + df[sig_col]

    return df


def _step5_age_calculations(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate Account Holder Age and Account Age."""
    anchor_date = _infer_report_date(df)

    if "DOB" in df.columns:
        df["DOB"] = pd.to_datetime(df["DOB"], errors="coerce", format="mixed")
        df["Account Holder Age"] = anchor_date.year - df["DOB"].dt.year

    if "Date Opened" in df.columns:
        df["Date Opened"] = pd.to_datetime(df["Date Opened"], errors="coerce", format="mixed")
        if "Date Closed" in df.columns:
            df["Date Closed"] = pd.to_datetime(df["Date Closed"], errors="coerce", format="mixed")
            end_date = df["Date Closed"].fillna(anchor_date)
        else:
            end_date = anchor_date
        df["Account Age"] = (end_date - df["Date Opened"]).dt.days / 365.25

    return df


def _step6_mail_response_grouping(df: pd.DataFrame) -> pd.DataFrame:
    """Count offers/responses and classify Response Grouping."""
    # Preserve pre-existing columns (some ODD files arrive with these populated)
    if "# of Offers" not in df.columns:
        mail_cols = [c for c in df.columns if c.endswith(" Mail")]
        if mail_cols:
            df["# of Offers"] = df[mail_cols].notna().sum(axis=1)
        else:
            df["# of Offers"] = 0

    if "# of Responses" not in df.columns:
        resp_cols = [c for c in df.columns if c.endswith(" Resp")]
        if resp_cols:
            resp_data = df[resp_cols].replace("NU 1-4", pd.NA)
            df["# of Responses"] = resp_data.notna().sum(axis=1)
        else:
            df["# of Responses"] = 0

    # Response Grouping classification
    rg = pd.Series("check", index=df.index)
    rg[df["# of Offers"] == 0] = "No Offer"
    rg[(df["# of Offers"] > 0) & (df["# of Responses"] == 0)] = "Non-Responder"
    rg[(df["# of Offers"] == 1) & (df["# of Responses"] == 1)] = "SO-SR"
    rg[(df["# of Offers"] >= 2) & (df["# of Responses"] == 1)] = "MO-SR"
    rg[df["# of Responses"] >= 2] = "MR"
    df["Response Grouping"] = rg

    return df


def _step7_control_segmentation(df: pd.DataFrame) -> pd.DataFrame:
    """Create per-month Segmentation columns (Control/Responder/Non-Responder)."""
    resp_cols = [c for c in df.columns if c.endswith(" Resp")]
    response_segments = {"NU 5+", "TH-10", "TH-15", "TH-20", "TH-25"}

    for resp_col in resp_cols:
        mail_col = resp_col.replace(" Resp", " Mail")
        seg_col = resp_col.replace(" Resp", " Segmentation")

        if mail_col not in df.columns:
            continue

        conditions = [
            df[mail_col].isna(),
            df[mail_col].notna() & (~df[resp_col].isin(response_segments)),
            df[mail_col].notna() & df[resp_col].isin(response_segments),
        ]
        choices = ["Control", "Non-Responder", "Responder"]
        df[seg_col] = np.select(conditions, choices, default="Control")

    return df


def _categorize_swipes(monthly_avg: float) -> str:
    """Categorize monthly swipe average into activity tiers."""
    if monthly_avg < 1:
        return "Non-user"
    elif monthly_avg <= 5:
        return "1-5 Swipes"
    elif monthly_avg <= 10:
        return "6-10 Swipes"
    elif monthly_avg <= 15:
        return "11-15 Swipes"
    elif monthly_avg <= 20:
        return "16-20 Swipes"
    elif monthly_avg <= 25:
        return "21-25 Swipes"
    elif monthly_avg <= 40:
        return "26-40 Swipes"
    else:
        return "41+ Swipes"


# ---------------------------------------------------------------------------
# Format validation -- header-only checks for ARS formatting and ICS readiness
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class FormatStatus:
    """Result of checking whether an ODD file has required columns."""

    is_formatted: bool
    found_columns: tuple[str, ...]
    missing_columns: tuple[str, ...]
    checked_path: str


# Columns produced by format_odd() steps 3/5/6 that do NOT exist in raw ODD.
ARS_SIGNATURE_COLUMNS: tuple[str, ...] = (
    "Total Spend",
    "Total Swipes",
    "SwipeCat12",
    "Account Holder Age",
    "Response Grouping",
)
ARS_SIGNATURE_THRESHOLD = 3  # 3 of 5 handles edge cases (missing DOB, no Mail cols)

# Columns appended by the ICS append step.
# Some institutions use "ICS Source", others use just "Source".
ICS_REQUIRED_COLUMNS: tuple[str, ...] = (
    "ICS Account",
    "ICS Source",
)
ICS_ALTERNATE_COLUMNS: tuple[str, ...] = (
    "ICS Account",
    "Source",
)
ICS_REQUIRED_THRESHOLD = 2  # both must be present


def check_odd_formatted(path: str | Path) -> FormatStatus:
    """Check if an ODD file has been through ARS formatting (header-only read).

    Reads column headers without loading the full dataset. Checks for
    signature columns that ``format_odd()`` adds (Total Spend, SwipeCat12, etc.).

    Returns a ``FormatStatus`` with ``is_formatted=True`` when at least
    ``ARS_SIGNATURE_THRESHOLD`` of the ``ARS_SIGNATURE_COLUMNS`` are found.
    """
    return _check_columns(path, ARS_SIGNATURE_COLUMNS, ARS_SIGNATURE_THRESHOLD)


def check_ics_ready(path: str | Path) -> FormatStatus:
    """Check if an ODD file has ICS columns (header-only read).

    Accepts either ``("ICS Account", "ICS Source")`` or ``("ICS Account", "Source")``
    since institutions vary in column naming.
    """
    result = _check_columns(path, ICS_REQUIRED_COLUMNS, ICS_REQUIRED_THRESHOLD)
    if result.is_formatted:
        return result
    return _check_columns(path, ICS_ALTERNATE_COLUMNS, ICS_REQUIRED_THRESHOLD)


def _check_columns(
    path: str | Path,
    signature: tuple[str, ...],
    threshold: int,
) -> FormatStatus:
    """Shared implementation for column-presence checks."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    suffix = path.suffix.lower()
    if suffix not in (".xlsx", ".xls", ".csv"):
        raise ValueError(f"Unsupported file format: {suffix}")

    headers = _read_column_headers(path)
    header_set = set(headers)

    found = tuple(c for c in signature if c in header_set)
    missing = tuple(c for c in signature if c not in header_set)

    return FormatStatus(
        is_formatted=len(found) >= threshold,
        found_columns=found,
        missing_columns=missing,
        checked_path=str(path),
    )


def _read_column_headers(path: Path) -> list[str]:
    """Read column headers only, without loading data rows."""
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            return [str(c) for c in row if c is not None]
        finally:
            wb.close()

    if suffix == ".xls":
        df = pd.read_excel(path, nrows=0, engine="xlrd")
        return list(df.columns)

    # CSV
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        first_row = next(reader, [])
    return [c.strip() for c in first_row]
