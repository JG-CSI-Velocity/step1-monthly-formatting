"""Shared Excel formatting and writing helpers.

Extracts the Excel export logic that was duplicated across ARS modules
and breaks the circular import (pipeline <-> analysis modules).
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Formatting constants
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
ALT_ROW_FILL = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
THIN_BORDER = Border(bottom=Side(style="thin", color="D0D0D0"))


def save_to_excel(
    wb: Workbook,
    df: pd.DataFrame | dict[str, pd.DataFrame],
    sheet_name: str,
    analysis_title: str,
    key_metrics: dict[str, str] | None = None,
) -> None:
    """Write a DataFrame (or dict of DataFrames) to a formatted Excel sheet."""
    sheet_title = _sanitize_sheet_title(sheet_name)

    if isinstance(df, dict):
        for sub_name, sub_df in df.items():
            sub_title = _sanitize_sheet_title(f"{sheet_title}-{sub_name}")
            _write_sheet(wb, sub_df, sub_title, analysis_title, key_metrics)
    else:
        _write_sheet(wb, df, sheet_title, analysis_title, key_metrics)


def _write_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    sheet_title: str,
    analysis_title: str,
    key_metrics: dict[str, str] | None = None,
) -> None:
    """Write a single formatted DataFrame to a new sheet."""
    ws = wb.create_sheet(title=sheet_title)
    row = 1

    # Title row
    ws.cell(row=row, column=1, value=analysis_title).font = Font(
        name="Calibri", size=14, bold=True, color="2E4057"
    )
    row += 1

    # Key metrics row
    if key_metrics:
        col = 1
        for label, value in key_metrics.items():
            ws.cell(row=row, column=col, value=f"{label}: {value}").font = Font(
                name="Calibri", size=10, italic=True, color="666666"
            )
            col += 1
        row += 1

    row += 1  # blank spacer

    # Write DataFrame
    for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row + r_idx, column=c_idx, value=value)
            if r_idx == 0:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if r_idx % 2 == 0:
                    cell.fill = ALT_ROW_FILL

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)


def create_workbook(title: str = "Analysis Report") -> Workbook:
    """Create a new workbook with a Summary sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value=title).font = Font(
        name="Calibri", size=16, bold=True, color="2E4057"
    )
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}").font = Font(
        name="Calibri", size=10, color="888888"
    )
    return wb


def save_workbook(wb: Workbook, path: Path, max_retries: int = 3) -> None:
    """Save workbook with retry logic for network drives."""
    import time

    path.parent.mkdir(parents=True, exist_ok=True)
    for attempt in range(max_retries):
        try:
            wb.save(str(path))
            return
        except PermissionError:
            if attempt < max_retries - 1:
                time.sleep(1)
            else:
                raise


def _sanitize_sheet_title(title: str) -> str:
    """Clean sheet title to meet Excel requirements (max 31 chars, no special chars)."""
    for char in r"[]:*?/\\":
        title = title.replace(char, "")
    return title[:31]
